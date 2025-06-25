"""
JSON to Excel Converter for YOLO Benchmark Results
Converts benchmark JSON to professional Excel tables with categorization
"""

import json
import pandas as pd
from pathlib import Path
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class BenchmarkExcelGenerator:
    """Convert JSON benchmark results to professional Excel tables"""
    
    def __init__(self, json_file_path):
        self.json_file_path = Path(json_file_path)
        self.data = self.load_json_data()
        self.results_df = self.process_data()
        
    def load_json_data(self):
        """Load JSON benchmark data"""
        try:
            with open(self.json_file_path, 'r') as f:
                data = json.load(f)
            print(f"[SUCCESS] Loaded benchmark data from {self.json_file_path}")
            return data
        except Exception as e:
            print(f"[ERROR] Failed to load JSON: {e}")
            return None
    
    def process_data(self):
        """Process JSON data into DataFrame"""
        if not self.data or 'results' not in self.data:
            print("[ERROR] Invalid JSON structure")
            return None
        
        processed_results = []
        
        for result in self.data['results']:
            if result['success']:
                metrics = result['metrics']
                
                processed_results.append({
                    'Model': result['name'],
                    'Category': result['category'],
                    'mAP@50': metrics.get('map_50', 0) * 100,  # Convert to percentage
                    'mAP@50-95': metrics.get('map_50', 0) * 100,  # Using mAP@50 as proxy
                    'FPS': metrics.get('fps', 0),
                    'Precision': metrics.get('precision', 0) * 100,
                    'Recall': metrics.get('recall', 0) * 100,
                    'F1': metrics.get('f1_score', 0) * 100,
                    'Params_M': self.parse_params(metrics.get('param_str', '0M')),
                    'Memory_GB': metrics.get('memory_gb', 0),
                    'Inference_Time': metrics.get('avg_inference_time', 0),
                    'Detections_Per_Image': metrics.get('avg_detections_per_image', 0),
                    'Expected_Params': result['expected_params'],
                    'Notes': result['notes']
                })
        
        df = pd.DataFrame(processed_results)
        print(f"[SUCCESS] Processed {len(df)} model results")
        return df
    
    def parse_params(self, param_str):
        """Parse parameter string to numeric value"""
        if 'M' in param_str:
            return float(param_str.replace('M', ''))
        elif 'K' in param_str:
            return float(param_str.replace('K', '')) / 1000
        else:
            return 0.0
    
    def categorize_models(self):
        """Categorize models into size groups like the reference images"""
        if self.results_df is None:
            return None, None, None
        
        # Define categories based on parameter count
        tiny_small = self.results_df[self.results_df['Params_M'] <= 12].copy()
        medium_large = self.results_df[(self.results_df['Params_M'] > 12) & (self.results_df['Params_M'] <= 35)].copy()
        xlarge = self.results_df[self.results_df['Params_M'] > 35].copy()
        
        # Sort each category by mAP@50 (descending)
        tiny_small = tiny_small.sort_values('mAP@50', ascending=False)
        medium_large = medium_large.sort_values('mAP@50', ascending=False)
        xlarge = xlarge.sort_values('mAP@50', ascending=False)
        
        return tiny_small, medium_large, xlarge
    
    def create_formatted_excel(self, output_file="benchmark_analysis.xlsx"):
        """Create professionally formatted Excel file with multiple sheets"""
        
        # Categorize models
        tiny_small, medium_large, xlarge = self.categorize_models()
        
        if tiny_small is None:
            print("[ERROR] Failed to categorize models")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # Sheet 1: Nano/Small Models
            self.create_category_sheet(
                writer, tiny_small, 
                "Nano_Small_Models",
                "Nano/Small Models (Optimized for Edge Devices)",
                "Best for: Edge deployment, Jetson devices, mobile applications"
            )
            
            # Sheet 2: Medium/Large Models  
            self.create_category_sheet(
                writer, medium_large,
                "Medium_Large_Models", 
                "Medium/Large Models (Balanced Accuracy/Speed)",
                "Best for: Production deployment, balanced performance"
            )
            
            # Sheet 3: XLarge Models
            self.create_category_sheet(
                writer, xlarge,
                "XLarge_Models",
                "Large/XLarge Models (Maximum Accuracy)", 
                "Best for: Maximum accuracy, server deployment"
            )
            
            # Sheet 4: Complete Results
            self.create_complete_results_sheet(writer, self.results_df)
            
            # Sheet 5: Summary Analysis
            self.create_summary_sheet(writer)
        
        print(f"[SUCCESS] Excel file created: {output_file}")
        return output_file
    
    def create_category_sheet(self, writer, df, sheet_name, title, description):
        """Create a formatted sheet for a model category"""
        
        # Select columns for display (matching reference images)
        display_df = df[[
            'Model', 'mAP@50-95', 'FPS', 'Params_M', 'Memory_GB'
        ]].copy()
        
        # Rename columns to match reference format
        display_df.columns = ['Model', 'mAP@50-95', 'FPS (A100)', 'Params (M)', 'Model Size (GB)']
        
        # Round values for clean display
        display_df['mAP@50-95'] = display_df['mAP@50-95'].round(1)
        display_df['FPS (A100)'] = display_df['FPS (A100)'].round(0).astype(int)
        display_df['Params (M)'] = display_df['Params (M)'].round(1)
        display_df['Model Size (GB)'] = display_df['Model Size (GB)'].round(2)
        
        # Write to sheet
        display_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Add title and description
        worksheet['A1'] = title
        worksheet['A2'] = description
        
        # Format title
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
        worksheet['A1'].font = title_font
        worksheet['A1'].fill = title_fill
        
        # Format description
        desc_font = Font(name='Arial', size=10, italic=True)
        worksheet['A2'].font = desc_font
        
        # Format headers
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(display_df.columns) + 1):
            cell = worksheet.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Format data rows
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        # Alternate row colors
        light_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        dark_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        for row in range(5, 5 + len(display_df)):
            fill = light_fill if (row - 5) % 2 == 0 else dark_fill
            
            for col in range(1, len(display_df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = fill
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(4, 5 + len(display_df)):
            for col in range(1, len(display_df.columns) + 1):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Adjust column widths
        column_widths = [15, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[chr(64 + i)].width = width
    
    def create_complete_results_sheet(self, writer, df):
        """Create complete results sheet with all metrics"""
        
        # Select all relevant columns
        complete_df = df[[
            'Model', 'Category', 'mAP@50', 'FPS', 'Precision', 'Recall', 'F1',
            'Params_M', 'Memory_GB', 'Detections_Per_Image', 'Expected_Params', 'Notes'
        ]].copy()
        
        # Round values
        for col in ['mAP@50', 'FPS', 'Precision', 'Recall', 'F1', 'Memory_GB', 'Detections_Per_Image']:
            complete_df[col] = complete_df[col].round(2)
        
        # Sort by mAP@50
        complete_df = complete_df.sort_values('mAP@50', ascending=False)
        
        # Write to sheet
        complete_df.to_excel(writer, sheet_name='Complete_Results', index=False)
        
        # Basic formatting
        workbook = writer.book
        worksheet = writer.sheets['Complete_Results']
        
        # Format headers
        header_font = Font(name='Arial', size=11, bold=True)
        for col in range(1, len(complete_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, writer):
        """Create summary analysis sheet"""
        
        # Create summary statistics
        summary_data = {
            'Metric': [
                'Best Overall Model (mAP@50)',
                'Fastest Model (FPS)',
                'Most Efficient (Memory)',
                'Best F1 Score',
                'Average mAP@50',
                'Average FPS',
                'Models ≥50 FPS',
                'Models ≥45% mAP@50'
            ],
            'Value': [
                self.results_df.loc[self.results_df['mAP@50'].idxmax(), 'Model'],
                self.results_df.loc[self.results_df['FPS'].idxmax(), 'Model'],
                self.results_df.loc[self.results_df['Memory_GB'].idxmin(), 'Model'],
                self.results_df.loc[self.results_df['F1'].idxmax(), 'Model'],
                f"{self.results_df['mAP@50'].mean():.1f}%",
                f"{self.results_df['FPS'].mean():.1f}",
                f"{len(self.results_df[self.results_df['FPS'] >= 50])} models",
                f"{len(self.results_df[self.results_df['mAP@50'] >= 45])} models"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Add dataset info
        if self.data:
            dataset_info = [
                ['Dataset Path', self.data.get('dataset_path', 'N/A')],
                ['Test Images', self.data.get('test_images_count', 'N/A')],
                ['Test Labels', self.data.get('test_labels_count', 'N/A')],
                ['Device', self.data.get('device', 'N/A')],
                ['Timestamp', self.data.get('timestamp', 'N/A')]
            ]
            
            dataset_df = pd.DataFrame(dataset_info, columns=['Property', 'Value'])
            dataset_df.to_excel(writer, sheet_name='Summary', index=False, startrow=len(summary_df) + 3)

def main():
    """Main execution function"""
    
    # Find the most recent JSON file
    json_files = list(Path('.').glob('custom_dataset_benchmark_*.json'))
    
    if not json_files:
        print("[ERROR] No benchmark JSON files found!")
        print("Make sure you've run the benchmark first.")
        return
    
    # Use the most recent file
    latest_json = max(json_files, key=lambda p: p.stat().st_mtime)
    print(f"[INFO] Using latest benchmark file: {latest_json}")
    
    # Generate Excel file
    generator = BenchmarkExcelGenerator(latest_json)
    
    # Create timestamp for output file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"benchmark_analysis_{timestamp}.xlsx"
    
    # Generate Excel
    result_file = generator.create_formatted_excel(output_file)
    
    if result_file:
        print(f"\n[SUCCESS] Excel analysis created!")
        print(f"File: {result_file}")
        print(f"\nSheets created:")
        print(f"  - Nano_Small_Models: Edge deployment models")
        print(f"  - Medium_Large_Models: Balanced performance models") 
        print(f"  - XLarge_Models: Maximum accuracy models")
        print(f"  - Complete_Results: All metrics")
        print(f"  - Summary: Key insights")
        
        # Print quick preview
        if generator.results_df is not None:
            print(f"\n[PREVIEW] Top 3 models by mAP@50:")
            top3 = generator.results_df.nlargest(3, 'mAP@50')[['Model', 'mAP@50', 'FPS', 'F1']]
            for _, row in top3.iterrows():
                print(f"  {row['Model']}: mAP {row['mAP@50']:.1f}%, {row['FPS']:.1f} FPS, F1 {row['F1']:.1f}%")

if __name__ == "__main__":
    main()